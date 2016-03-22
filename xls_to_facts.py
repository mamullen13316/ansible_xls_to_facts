#!/Library/Frameworks/Python.framework/Versions/2.7/bin/python


"""
     Copyright (c) 2015 World Wide Technology, Inc.
     All rights reserved.

     Revision history:
     26 Sept 2015  |  1.0 - initial release


"""

DOCUMENTATION = '''
---
module: XLS_to_facts.py
author: Joel W. King, World Wide Technology
version_added: "1.0"
short_description: Read a XLS file and output Ansible facts
description:
    - Read the XLS file specified and output Ansible facts in the form of a list with each
      element in the list as a dictionary using the column header as the key and the contents
      of the cell as the value.
 
requirements:
    - None

options:
    src:
        description:
            - The XLS formatted input file
        required: true

    
'''

EXAMPLES = '''

    Running the module

      ansible  localhost  -m XLS_to_facts -a "src=/tmp/TonyA.XLS"

    In a role configuration, given a group and host entry

      [asante]
      NEX-3048-E.sandbox.wwtatc.local  ansible_connection=local ansible_ssh_user=kingjoe hostname=13leafzn02-rp01
      #

      $ cat asante.yml
      ---
      - name: Test  Role
        hosts: asante

        roles:
          - {role: excel_nxos, debug: on}


      $ ansible-playbook asante.yml --ask-vault

'''

import openpyxl

# ---------------------------------------------------------------------------
# read_xls_dict
# ---------------------------------------------------------------------------

def read_xls_dict(input_file):
    "Read the XLS file and return as Ansible facts"
    result = {"ansible_facts":{}}
    spreadsheet = {}
    try:
        wb = openpyxl.load_workbook(input_file)
        for sheet in wb.get_sheet_names():
            ansible_sheet_name = 'spreadsheet_' + sheet
            spreadsheet[ansible_sheet_name] = []
            current_sheet = wb.get_sheet_by_name(sheet)
            dict_keys = []
            for c in range(1,current_sheet.max_column + 1):
                dict_keys.append(current_sheet.cell(row=1,column=c).value)
            for r in range (2,current_sheet.max_row + 1):
                temp_dict = {}
                for c in range(1,current_sheet.max_column + 1):
                    temp_dict[dict_keys[c-1]] = current_sheet.cell(row=r,column=c).value
                spreadsheet[ansible_sheet_name].append(temp_dict)
    except IOError:
        return (1, "IOError on input file:%s" % input_file)

    result["ansible_facts"] = spreadsheet
    return (0, result)

# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

def main():
    " "
    module = AnsibleModule(argument_spec = dict(
             src = dict(required=True)
             ),
             check_invalid_arguments=False,
             add_file_common_args=True)

    code, response = read_xls_dict(module.params["src"])
    if code == 1:
        module.fail_json(msg=response)
    else:
        module.exit_json(**response)

    return code


from ansible.module_utils.basic import *
main()
#
